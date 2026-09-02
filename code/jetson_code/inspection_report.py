from __future__ import annotations

import os
from datetime import datetime
from numbers import Integral, Real
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAME = "점검 결과"
TOP_CRACK_SHEET_NAME = "상단 크랙 결과"
SETTINGS_SHEET_NAME = "검사 설정"
PIPELINE_VERSION = 19
CAPTURE_DETECTOR_PREFIX = "deeplabv3plus-tensorrt/teacher/"
REALTIME_DETECTOR_PREFIX = "deeplabv3plus-tensorrt/student/"
OPTIMIZED_REALTIME_DETECTOR_PREFIX = (
    "deeplabv3plus-tensorrt/student/optimized-compact-v2/"
)
CUDA_ARGMAX_REALTIME_DETECTOR_SUFFIX = "/gpu-argmax"
HRSEGNET_CAPTURE_CRACK_DETECTOR_PREFIX = (
    "hrsegnet-b32-tensorrt/capture/crack/"
)
REALTIME_CRACK_DETECTOR_PREFIX = "bgcrack-tensorrt/realtime/"
HRSEGNET_REALTIME_CRACK_DETECTOR_PREFIX = (
    "hrsegnet-b32-tensorrt/realtime/crack/"
)
MULTITASK_REALTIME_DETECTOR_PREFIX = (
    "multitask-segmentation-tensorrt/realtime/rust/"
)
MULTITASK_REALTIME_CRACK_DETECTOR_PREFIX = (
    "multitask-segmentation-tensorrt/realtime/crack/"
)
CAPTURE_MODEL_ARCHITECTURE = "DeepLabV3+ ResNet101, output_stride=8"
REALTIME_MODEL_ARCHITECTURE = (
    "distilled DeepLabV3+ MobileNetV2, output_stride=8"
)
CAPTURE_INPUT_PREPROCESSING = (
    "full camera frame 1280x720; no resize/padding/letterbox; OpenCV BGR; "
    "float32 0..255; no scaling or mean/std"
)
REALTIME_INPUT_PREPROCESSING = (
    "full-width top ROI y=0:240 (1280x240); no resize/padding; "
    "BGR->RGB; 0..1; ImageNet mean/std"
)
CAPTURE_MODEL_IO_CONTRACT = (
    "images float32 [1,3,720,1280] -> logits float32 [1,4,720,1280]"
)
REALTIME_MODEL_IO_CONTRACT = (
    "images float32 [1,3,240,1280] -> logits float32 [1,4,240,1280]"
)
CUDA_ARGMAX_REALTIME_MODEL_IO_CONTRACT = (
    "TensorRT images float32 [1,3,240,1280] -> logits float32 "
    "[1,4,240,1280] kept on DEVICE; same-stream CUDA finite check and "
    "first-tie argmax -> class_map uint8 [240,1280] plus invalid uint32 [1]; "
    "raw logits D2H forbidden"
)
OPTIMIZED_REALTIME_MODEL_IO_CONTRACT = (
    "images float32 [1,3,240,1280] -> rust_class_map uint8 "
    "[1,240,1280], rust_logits_not_nan uint8 [1] (must equal 1), "
    "rust_logits_finite_abs uint8 [1] (must equal 1); CPU derives "
    "class counts and Poor/Severe control from the transferred map; "
    "raw 4-channel D2H forbidden"
)
HRSEGNET_CAPTURE_CRACK_MODEL_ARCHITECTURE = (
    "original HrSegNet-B32 CrackSeg9k semantic segmentation capture model"
)
HRSEGNET_CAPTURE_CRACK_INPUT_PREPROCESSING = (
    "full camera frame 1280x720; no resize/padding; BGR->RGB; "
    "float32 /127.5-1.0"
)
HRSEGNET_CAPTURE_CRACK_MODEL_IO_CONTRACT = (
    "images float32 [1,3,720,1280] -> crack_logits float32 "
    "[1,2,720,1280]; finite class-1 minus class-0 logit margin; "
    "probability threshold converted to a logit-margin threshold; "
    "CPU 8-connected component filtering"
)
REALTIME_CRACK_INPUT_PREPROCESSING = (
    "full-width ROI y=112:240 (1280x128); no resize/padding; BGR->RGB; "
    "float32 /127.5-1.0"
)
REALTIME_CRACK_MODEL_IO_CONTRACT = (
    "images float32 [1,3,128,1280] -> crack_probability float32 "
    "[1,1,128,1280]; no sigmoid or spatial remap"
)
HRSEGNET_REALTIME_CRACK_INPUT_PREPROCESSING = (
    "full-width ROI y=112:240 (1280x128); no resize/padding; BGR->RGB; "
    "float32 /127.5-1.0"
)
HRSEGNET_REALTIME_CRACK_MODEL_IO_CONTRACT = (
    "images float32 [1,3,128,1280] -> crack_logits float32 "
    "[1,2,128,1280]; finite class-1 minus class-0 logit margin; "
    "probability threshold converted to a logit-margin threshold; "
    "CPU 8-connected component filtering"
)
MULTITASK_REALTIME_MODEL_ARCHITECTURE = (
    "shared MobileNet realtime rust/crack semantic segmentation; "
    "GPU argmax/threshold postprocessor"
)
MULTITASK_REALTIME_INPUT_PREPROCESSING = (
    "full-width top ROI y=0:240 (1280x240); no resize/padding; "
    "BGR->RGB; 0..1; ImageNet mean/std; crack valid rows=112:240"
)
MULTITASK_REALTIME_MODEL_IO_CONTRACT = (
    "images float32 [1,3,240,1280] -> rust_class_map uint8 "
    "[1,240,1280], rust_class_counts int32 [4], rust_poor_severe "
    "uint8 [1], crack_candidate_map uint8 [1,128,1280], "
    "crack_candidate_pixels int32 [1], crack_probability_threshold "
    "float32 [1], multitask_outputs_finite uint8 [1] (must equal 1); "
    "raw 5-channel D2H forbidden"
)
MULTITASK_REALTIME_CRACK_INPUT_PREPROCESSING = (
    "shared normalized images input; crack valid rows y=112:240 from the "
    "1280x240 top ROI; GPU logit threshold"
)
MULTITASK_REALTIME_CRACK_MODEL_IO_CONTRACT = (
    "same optimized multitask plan; crack_candidate_map uint8 "
    "[1,128,1280] plus count and embedded probability-threshold scalars"
)
REPORT_RESULT_POLICY = (
    "SIDE capture rows contain full-frame corrosion and capture HrSegNet candidates; "
    "TOP crack-only captures use the separate top crack sheet; realtime rust/crack "
    "engines are control provenance only"
)
CLASS_ORDER = "Good,Fair,Poor,Severe"
HEADER_ROW = 6
FIRST_DATA_ROW = HEADER_ROW + 1
INITIAL_PHASE = "initial"
RESCAN_PHASE = "rescan"
MANUAL_PHASE = "manual"
RAIL_SECTION_TARGET = 4

PHASE_LABELS = {
    INITIAL_PHASE: "초기 검사",
    RESCAN_PHASE: "재검사",
    MANUAL_PHASE: "수동 검사",
}
PHASE_BY_LABEL = {label: phase for phase, label in PHASE_LABELS.items()}
PHASE_TRIGGERS = {
    INITIAL_PHASE: "CAMERA_CAPTURE",
    RESCAN_PHASE: "CAMERA_CAPTURE",
    MANUAL_PHASE: "manual",
}

HEADERS = (
    "번호",
    "검사 단계",
    "Rail Section 번호",
    "촬영시각",
    "사진 경로",
    "트리거",
    "검출기",
    "부식 픽셀",
    "검사 픽셀",
    "사진 부식률",
    "레일 구간",
    "균열 분석 상태",
    "균열 검출기",
    "균열 후보 감지",
    "균열 픽셀",
    "균열 검사 픽셀",
    "사진 균열률",
)

TOP_CRACK_HEADERS = (
    "번호",
    "검사 단계",
    "Rail Section 번호",
    "촬영시각",
    "원본 사진 경로",
    "분석 사진 경로",
    "트리거",
    "카메라 역할",
    "균열 분석 상태",
    "균열 검출기",
    "균열 후보 감지",
    "균열 픽셀",
    "균열 검사 픽셀",
    "사진 균열률",
    "레일 구간",
)
TOP_CRACK_HEADER_ROW = 1
TOP_CRACK_FIRST_DATA_ROW = TOP_CRACK_HEADER_ROW + 1

THIN_GRAY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


class InspectionWorkbook:
    """Append-only XLSX report for SIDE corrosion/crack and TOP crack captures."""

    def __init__(
        self,
        path: str | Path,
        capture_model_filename: str,
        capture_model_sha256: str,
        realtime_model_filename: str,
        realtime_model_sha256: str,
        capture_detector: str,
        realtime_detector: str,
        crack_model_filename: str | None = None,
        crack_model_sha256: str | None = None,
        crack_detector: str | None = None,
        crack_probability_threshold: float = 0.55,
        capture_crack_min_component_pixels: int = 20,
        realtime_crack_model_filename: str | None = None,
        realtime_crack_model_sha256: str | None = None,
        realtime_crack_detector: str | None = None,
        realtime_crack_probability_threshold: float = 0.55,
        realtime_crack_min_component_pixels: int = 20,
    ) -> None:
        self.path = Path(path).expanduser()
        self.capture_model_filename = Path(capture_model_filename).name
        self.realtime_model_filename = Path(realtime_model_filename).name
        if not self.capture_model_filename or not self.realtime_model_filename:
            raise ValueError("TensorRT model filenames must not be empty.")
        self.capture_model_sha256 = self._validate_model_sha256(
            capture_model_sha256
        )
        self.realtime_model_sha256 = self._validate_model_sha256(
            realtime_model_sha256
        )
        self.capture_detector = str(capture_detector)
        self.realtime_detector = str(realtime_detector)
        if not self.capture_detector.startswith(CAPTURE_DETECTOR_PREFIX):
            raise ValueError("capture_detector must identify the TensorRT teacher.")
        self.realtime_optimized_enabled = self.realtime_detector.startswith(
            OPTIMIZED_REALTIME_DETECTOR_PREFIX
        )
        self.realtime_cuda_argmax_enabled = (
            self.realtime_detector.startswith(REALTIME_DETECTOR_PREFIX)
            and self.realtime_detector.endswith(
                CUDA_ARGMAX_REALTIME_DETECTOR_SUFFIX
            )
        )
        self.realtime_multitask_enabled = self.realtime_detector.startswith(
            MULTITASK_REALTIME_DETECTOR_PREFIX
        )
        if (
            not self.realtime_optimized_enabled
            and not self.realtime_multitask_enabled
            and not self.realtime_detector.startswith(REALTIME_DETECTOR_PREFIX)
        ):
            raise ValueError(
                "realtime_detector must identify the TensorRT student or the "
                "optimized multitask rust role."
            )
        crack_values = (
            crack_model_filename,
            crack_model_sha256,
            crack_detector,
        )
        if any(value is not None for value in crack_values) and not all(
            value is not None for value in crack_values
        ):
            raise ValueError(
                "Crack model filename, SHA-256, and detector must be provided together."
            )
        self.crack_enabled = all(value is not None for value in crack_values)
        if self.crack_enabled:
            self.crack_model_filename = Path(str(crack_model_filename)).name
            if not self.crack_model_filename:
                raise ValueError("Crack TensorRT model filename must not be empty.")
            self.crack_model_sha256 = self._validate_model_sha256(
                str(crack_model_sha256)
            )
            self.crack_detector = str(crack_detector)
            self.capture_crack_hrsegnet_enabled = self.crack_detector.startswith(
                HRSEGNET_CAPTURE_CRACK_DETECTOR_PREFIX
            )
            if not self.capture_crack_hrsegnet_enabled:
                raise ValueError(
                    "crack_detector must identify the capture HrSegNet TensorRT model."
                )
        else:
            self.crack_model_filename = "N/A"
            self.crack_model_sha256 = "N/A"
            self.crack_detector = "disabled"
            self.capture_crack_hrsegnet_enabled = False
        realtime_crack_values = (
            realtime_crack_model_filename,
            realtime_crack_model_sha256,
            realtime_crack_detector,
        )
        if any(value is not None for value in realtime_crack_values) and not all(
            value is not None for value in realtime_crack_values
        ):
            raise ValueError(
                "Realtime crack model filename, SHA-256, and detector must be "
                "provided together."
            )
        self.realtime_crack_enabled = all(
            value is not None for value in realtime_crack_values
        )
        if self.realtime_crack_enabled != self.crack_enabled:
            raise ValueError(
                "Capture and realtime crack provenance must both be enabled or "
                "both be disabled."
            )
        if self.realtime_crack_enabled:
            self.realtime_crack_model_filename = Path(
                str(realtime_crack_model_filename)
            ).name
            if not self.realtime_crack_model_filename:
                raise ValueError(
                    "Realtime crack TensorRT model filename must not be empty."
                )
            self.realtime_crack_model_sha256 = self._validate_model_sha256(
                str(realtime_crack_model_sha256)
            )
            self.realtime_crack_detector = str(realtime_crack_detector)
            realtime_crack_is_multitask = self.realtime_crack_detector.startswith(
                MULTITASK_REALTIME_CRACK_DETECTOR_PREFIX
            )
            realtime_crack_is_hrsegnet = self.realtime_crack_detector.startswith(
                HRSEGNET_REALTIME_CRACK_DETECTOR_PREFIX
            )
            self.realtime_crack_multitask_enabled = realtime_crack_is_multitask
            self.realtime_crack_hrsegnet_enabled = realtime_crack_is_hrsegnet
            if (
                not realtime_crack_is_multitask
                and not realtime_crack_is_hrsegnet
                and not self.realtime_crack_detector.startswith(
                    REALTIME_CRACK_DETECTOR_PREFIX
                )
            ):
                raise ValueError(
                    "realtime_crack_detector must identify the realtime BGCrack "
                    "model, HrSegNet crack role, or optimized multitask crack role."
                )
            if self.realtime_multitask_enabled and not realtime_crack_is_multitask:
                raise ValueError(
                    "Optimized multitask rust provenance requires optimized "
                    "multitask crack provenance."
                )
            if self.realtime_multitask_enabled and (
                self.realtime_model_filename != self.realtime_crack_model_filename
                or self.realtime_model_sha256 != self.realtime_crack_model_sha256
            ):
                raise ValueError(
                    "Optimized multitask rust and crack roles must reference the "
                    "same TensorRT plan and SHA-256."
                )
            if realtime_crack_is_multitask and not self.realtime_multitask_enabled and (
                self.realtime_model_filename == self.realtime_crack_model_filename
                or self.realtime_model_sha256 == self.realtime_crack_model_sha256
            ):
                raise ValueError(
                    "Hybrid realtime rust and multitask crack roles must reference "
                    "different TensorRT plans and SHA-256 values."
                )
        else:
            self.realtime_crack_model_filename = "N/A"
            self.realtime_crack_model_sha256 = "N/A"
            self.realtime_crack_detector = "disabled"
            self.realtime_crack_multitask_enabled = False
            self.realtime_crack_hrsegnet_enabled = False
        if (
            not isinstance(crack_probability_threshold, Real)
            or isinstance(crack_probability_threshold, bool)
            or not 0.0 < float(crack_probability_threshold) < 1.0
        ):
            raise ValueError(
                "crack_probability_threshold must be between zero and one."
            )
        if (
            not isinstance(capture_crack_min_component_pixels, Integral)
            or isinstance(capture_crack_min_component_pixels, bool)
            or int(capture_crack_min_component_pixels) <= 0
        ):
            raise ValueError(
                "capture_crack_min_component_pixels must be a positive integer."
            )
        self.crack_probability_threshold = float(crack_probability_threshold)
        self.capture_crack_min_component_pixels = int(
            capture_crack_min_component_pixels
        )
        if (
            not isinstance(realtime_crack_probability_threshold, Real)
            or isinstance(realtime_crack_probability_threshold, bool)
            or not 0.0 < float(realtime_crack_probability_threshold) < 1.0
        ):
            raise ValueError(
                "realtime_crack_probability_threshold must be between zero and one."
            )
        self.realtime_crack_probability_threshold = float(
            realtime_crack_probability_threshold
        )
        if (
            not isinstance(realtime_crack_min_component_pixels, Integral)
            or isinstance(realtime_crack_min_component_pixels, bool)
            or int(realtime_crack_min_component_pixels) <= 0
        ):
            raise ValueError(
                "realtime_crack_min_component_pixels must be a positive integer."
            )
        self.realtime_crack_min_component_pixels = int(
            realtime_crack_min_component_pixels
        )
        if self.realtime_multitask_enabled:
            self.realtime_model_architecture = MULTITASK_REALTIME_MODEL_ARCHITECTURE
            self.realtime_input_preprocessing = MULTITASK_REALTIME_INPUT_PREPROCESSING
            self.realtime_model_io_contract = MULTITASK_REALTIME_MODEL_IO_CONTRACT
        elif self.realtime_cuda_argmax_enabled:
            self.realtime_model_architecture = REALTIME_MODEL_ARCHITECTURE
            self.realtime_input_preprocessing = REALTIME_INPUT_PREPROCESSING
            self.realtime_model_io_contract = CUDA_ARGMAX_REALTIME_MODEL_IO_CONTRACT
        elif self.realtime_optimized_enabled:
            self.realtime_model_architecture = REALTIME_MODEL_ARCHITECTURE
            self.realtime_input_preprocessing = REALTIME_INPUT_PREPROCESSING
            self.realtime_model_io_contract = OPTIMIZED_REALTIME_MODEL_IO_CONTRACT
        else:
            self.realtime_model_architecture = REALTIME_MODEL_ARCHITECTURE
            self.realtime_input_preprocessing = REALTIME_INPUT_PREPROCESSING
            self.realtime_model_io_contract = REALTIME_MODEL_IO_CONTRACT
        if self.realtime_crack_multitask_enabled:
            self.realtime_crack_input_preprocessing = (
                MULTITASK_REALTIME_CRACK_INPUT_PREPROCESSING
            )
            self.realtime_crack_model_io_contract = (
                MULTITASK_REALTIME_CRACK_MODEL_IO_CONTRACT
            )
        elif self.realtime_crack_hrsegnet_enabled:
            self.realtime_crack_input_preprocessing = (
                HRSEGNET_REALTIME_CRACK_INPUT_PREPROCESSING
            )
            self.realtime_crack_model_io_contract = (
                HRSEGNET_REALTIME_CRACK_MODEL_IO_CONTRACT
            )
        else:
            self.realtime_crack_input_preprocessing = REALTIME_CRACK_INPUT_PREPROCESSING
            self.realtime_crack_model_io_contract = REALTIME_CRACK_MODEL_IO_CONTRACT
        if self.capture_crack_hrsegnet_enabled:
            self.capture_crack_model_architecture = (
                HRSEGNET_CAPTURE_CRACK_MODEL_ARCHITECTURE
            )
            self.capture_crack_input_preprocessing = (
                HRSEGNET_CAPTURE_CRACK_INPUT_PREPROCESSING
            )
            self.capture_crack_model_io_contract = (
                HRSEGNET_CAPTURE_CRACK_MODEL_IO_CONTRACT
            )
        else:
            self.capture_crack_model_architecture = "disabled"
            self.capture_crack_input_preprocessing = "N/A"
            self.capture_crack_model_io_contract = "N/A"
        if self.path.suffix.lower() != ".xlsx":
            raise ValueError("Inspection report path must use the .xlsx extension.")
        if self.path.exists():
            self.workbook = load_workbook(self.path)
            if not {
                SHEET_NAME,
                TOP_CRACK_SHEET_NAME,
                SETTINGS_SHEET_NAME,
            }.issubset(self.workbook.sheetnames):
                raise ValueError(
                    "Existing workbook does not contain the required inspection sheets."
                )
            self.sheet = self.workbook[SHEET_NAME]
            self.top_crack_sheet = self.workbook[TOP_CRACK_SHEET_NAME]
            self.settings_sheet = self.workbook[SETTINGS_SHEET_NAME]
            self._validate_settings()
            self._validate_existing_sheet()
            self._validate_top_crack_sheet()
            phase_counts = self._phase_counts()
            if phase_counts[INITIAL_PHASE] or phase_counts[RESCAN_PHASE]:
                raise ValueError(
                    "Report already contains automatic inspection rows; "
                    "use a new --report file for this run."
                )
        else:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = SHEET_NAME
            self._initialize_sheet(self.sheet)
            self.top_crack_sheet = self.workbook.create_sheet(
                TOP_CRACK_SHEET_NAME
            )
            self._initialize_top_crack_sheet(self.top_crack_sheet)
            self.settings_sheet = self.workbook.create_sheet(SETTINGS_SHEET_NAME)
            self._initialize_settings(self.settings_sheet)
            self._save()

    def append_capture(
        self,
        capture_path: str | Path,
        phase: str,
        trigger: str,
        detector: str,
        rust_pixels: int,
        inspected_pixels: int,
        captured_at: datetime | None = None,
        *,
        crack_status: str = "disabled",
        crack_detector: str | None = None,
        crack_detected: bool | None = None,
        crack_pixels: int | None = None,
        crack_inspected_pixels: int | None = None,
    ) -> float:
        """Append one capture and return its phase's pixel-weighted ratio."""

        rust_pixels, inspected_pixels = self._validate_pixels(
            rust_pixels, inspected_pixels
        )
        captured_at = self._normalize_captured_at(captured_at)
        phase = str(phase)
        if phase not in PHASE_LABELS:
            raise ValueError(f"Unknown inspection phase: {phase}")
        trigger = str(trigger)
        if trigger != PHASE_TRIGGERS[phase]:
            raise ValueError(f"Unexpected trigger {trigger!r} for phase {phase!r}.")
        detector = str(detector)
        if detector != self.capture_detector:
            raise ValueError(
                "Only results from this report's exact capture teacher can be added."
            )
        (
            crack_status,
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        ) = self._validate_crack_measurement(
            crack_status,
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        )

        phase_counts = self._phase_counts()
        if phase == INITIAL_PHASE:
            if phase_counts[RESCAN_PHASE] != 0:
                raise ValueError(
                    "Initial rail sections cannot be added after rescan starts."
                )
            if phase_counts[INITIAL_PHASE] >= RAIL_SECTION_TARGET:
                raise ValueError(
                    "The report already contains "
                    f"{RAIL_SECTION_TARGET} initial rail sections."
                )
        elif phase == RESCAN_PHASE:
            if phase_counts[INITIAL_PHASE] != RAIL_SECTION_TARGET:
                raise ValueError(
                    "Rescan requires exactly "
                    f"{RAIL_SECTION_TARGET} initial rail sections first."
                )
            if phase_counts[RESCAN_PHASE] >= RAIL_SECTION_TARGET:
                raise ValueError(
                    "The report already contains "
                    f"{RAIL_SECTION_TARGET} rescan rail sections."
                )

        row = self._next_data_row()
        sequence = row - FIRST_DATA_ROW + 1
        phase_sequence = phase_counts[phase] + 1
        zone_label = (
            f"{phase_sequence}번 레일 구간"
            if phase in (INITIAL_PHASE, RESCAN_PHASE)
            else "수동"
        )
        values = (
            sequence,
            PHASE_LABELS[phase],
            phase_sequence,
            captured_at,
            str(capture_path),
            trigger,
            detector,
            rust_pixels,
            inspected_pixels,
        )
        for column, value in enumerate(values, start=1):
            self.sheet.cell(row=row, column=column, value=value)
        self.sheet.cell(row=row, column=10, value=f"=IF(I{row}=0,0,H{row}/I{row})")
        crack_values = (
            zone_label,
            crack_status,
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        )
        for column, value in enumerate(crack_values, start=11):
            self.sheet.cell(row=row, column=column, value=value)
        if crack_status == "ready":
            self.sheet.cell(
                row=row,
                column=17,
                value=f"=IF(P{row}=0,0,O{row}/P{row})",
            )
        self._format_data_row(row)

        total_rust = sum(
            int(self.sheet.cell(data_row, 8).value or 0)
            for data_row in range(FIRST_DATA_ROW, row + 1)
            if self.sheet.cell(data_row, 2).value == PHASE_LABELS[phase]
        )
        total_inspected = sum(
            int(self.sheet.cell(data_row, 9).value or 0)
            for data_row in range(FIRST_DATA_ROW, row + 1)
            if self.sheet.cell(data_row, 2).value == PHASE_LABELS[phase]
        )
        overall_ratio = total_rust / total_inspected
        self._update_summary_formulas(row)
        self.sheet.auto_filter.ref = f"A{HEADER_ROW}:Q{row}"
        try:
            self._save()
        except Exception:
            self._reload()
            raise
        return overall_ratio

    def append_top_crack_capture(
        self,
        raw_capture_path: str | Path,
        capture_path: str | Path,
        phase: str,
        phase_sequence: int | None,
        trigger: str,
        crack_detector: str,
        crack_detected: bool,
        crack_pixels: int,
        crack_inspected_pixels: int,
        captured_at: datetime | None = None,
    ) -> float:
        """Append one TOP-camera crack result without altering SIDE rust totals."""

        captured_at = self._normalize_captured_at(captured_at)
        phase = str(phase)
        if phase not in PHASE_LABELS:
            raise ValueError(f"Unknown inspection phase: {phase}")
        trigger = str(trigger)
        if trigger != PHASE_TRIGGERS[phase]:
            raise ValueError(f"Unexpected trigger {trigger!r} for phase {phase!r}.")
        if phase in (INITIAL_PHASE, RESCAN_PHASE):
            if (
                not isinstance(phase_sequence, Integral)
                or isinstance(phase_sequence, bool)
                or not 1 <= int(phase_sequence) <= RAIL_SECTION_TARGET
            ):
                raise ValueError(
                    "Automatic top rail-section sequence must be from 1 to "
                    f"{RAIL_SECTION_TARGET}."
                )
            phase_sequence = int(phase_sequence)
        elif phase_sequence is not None:
            raise ValueError("Manual top rail-section sequence must be None.")

        (
            _status,
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        ) = self._validate_crack_measurement(
            "ready",
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        )

        if not self._matching_side_capture_exists(
            phase,
            phase_sequence,
            captured_at,
        ):
            raise ValueError(
                "Top crack capture requires its matching SIDE report row first."
            )
        top_phase_counts = self._top_crack_phase_counts()
        if phase == INITIAL_PHASE:
            if top_phase_counts[RESCAN_PHASE] != 0:
                raise ValueError(
                    "Initial top rail sections cannot follow top rescan rows."
                )
            expected_sequence = top_phase_counts[INITIAL_PHASE] + 1
            if phase_sequence != expected_sequence:
                raise ValueError(
                    f"Expected top initial rail section {expected_sequence}; "
                    f"got {phase_sequence}."
                )
        elif phase == RESCAN_PHASE:
            if top_phase_counts[INITIAL_PHASE] != RAIL_SECTION_TARGET:
                raise ValueError(
                    "Top rescan requires exactly "
                    f"{RAIL_SECTION_TARGET} initial top rail sections first."
                )
            expected_sequence = top_phase_counts[RESCAN_PHASE] + 1
            if phase_sequence != expected_sequence:
                raise ValueError(
                    f"Expected top rescan rail section {expected_sequence}; "
                    f"got {phase_sequence}."
                )

        row = self._next_top_crack_row()
        sequence = row - TOP_CRACK_FIRST_DATA_ROW + 1
        zone_label = (
            f"{phase_sequence}번 레일 구간"
            if phase in (INITIAL_PHASE, RESCAN_PHASE)
            else "수동"
        )
        values = (
            sequence,
            PHASE_LABELS[phase],
            phase_sequence,
            captured_at,
            str(raw_capture_path),
            str(capture_path),
            trigger,
            "top",
            "ready",
            crack_detector,
            crack_detected,
            crack_pixels,
            crack_inspected_pixels,
        )
        for column, value in enumerate(values, start=1):
            self.top_crack_sheet.cell(row=row, column=column, value=value)
        self.top_crack_sheet.cell(
            row=row,
            column=14,
            value=f"=IF(M{row}=0,0,L{row}/M{row})",
        )
        self.top_crack_sheet.cell(row=row, column=15, value=zone_label)
        self._format_top_crack_row(row)
        self.top_crack_sheet.auto_filter.ref = (
            f"A{TOP_CRACK_HEADER_ROW}:O{row}"
        )
        try:
            self._save()
        except Exception:
            self._reload()
            raise
        return crack_pixels / crack_inspected_pixels

    def _save(self) -> None:
        temporary_path = None
        try:
            with NamedTemporaryFile(
                prefix=f".{self.path.stem}_",
                suffix=".xlsx",
                dir=self.path.parent,
                delete=False,
            ) as temporary_file:
                temporary_path = Path(temporary_file.name)
            self.workbook.save(temporary_path)
            os.replace(temporary_path, self.path)
            temporary_path = None
        finally:
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)

    def _reload(self) -> None:
        self.workbook = load_workbook(self.path)
        self.sheet = self.workbook[SHEET_NAME]
        self.top_crack_sheet = self.workbook[TOP_CRACK_SHEET_NAME]
        self.settings_sheet = self.workbook[SETTINGS_SHEET_NAME]

    def _initialize_settings(self, sheet: Worksheet) -> None:
        sheet.append(("설정", "값"))
        sheet.append(("파이프라인 버전", PIPELINE_VERSION))
        sheet.append(("Rail Section 목표", RAIL_SECTION_TARGET))
        sheet.append(("캡처 모델 구조", CAPTURE_MODEL_ARCHITECTURE))
        sheet.append(("캡처 모델 파일", self.capture_model_filename))
        sheet.append(("캡처 모델 SHA-256", self.capture_model_sha256))
        sheet.append(("캡처 검출기", self.capture_detector))
        sheet.append(("캡처 입력 전처리", CAPTURE_INPUT_PREPROCESSING))
        sheet.append(("실시간 모델 구조", self.realtime_model_architecture))
        sheet.append(("실시간 모델 파일", self.realtime_model_filename))
        sheet.append(("실시간 모델 SHA-256", self.realtime_model_sha256))
        sheet.append(("실시간 검출기", self.realtime_detector))
        sheet.append(("실시간 입력 전처리", self.realtime_input_preprocessing))
        sheet.append(("캡처 모델 I/O 계약", CAPTURE_MODEL_IO_CONTRACT))
        sheet.append(("실시간 모델 I/O 계약", self.realtime_model_io_contract))
        sheet.append(("결과 기록 정책", REPORT_RESULT_POLICY))
        sheet.append(("클래스 순서", CLASS_ORDER))
        sheet.append(("부식 픽셀 정의", "Fair/Poor/Severe"))
        sheet.append(("균열 분석 상태", "enabled" if self.crack_enabled else "disabled"))
        sheet.append(("균열 모델 구조", self.capture_crack_model_architecture))
        sheet.append(("균열 모델 파일", self.crack_model_filename))
        sheet.append(("균열 모델 SHA-256", self.crack_model_sha256))
        sheet.append(("균열 검출기", self.crack_detector))
        sheet.append(("균열 입력 전처리", self.capture_crack_input_preprocessing))
        sheet.append(("균열 모델 I/O 계약", self.capture_crack_model_io_contract))
        sheet.append(("균열 확률 임계값", self.crack_probability_threshold))
        sheet.append(
            (
                "캡처 균열 최소 연결성분 픽셀",
                self.capture_crack_min_component_pixels,
            )
        )
        sheet.append(("실시간 균열 모델 파일", self.realtime_crack_model_filename))
        sheet.append(("실시간 균열 모델 SHA-256", self.realtime_crack_model_sha256))
        sheet.append(("실시간 균열 검출기", self.realtime_crack_detector))
        sheet.append(
            ("실시간 균열 입력 전처리", self.realtime_crack_input_preprocessing)
        )
        sheet.append(
            ("실시간 균열 모델 I/O 계약", self.realtime_crack_model_io_contract)
        )
        sheet.append(
            ("실시간 균열 확률 임계값", self.realtime_crack_probability_threshold)
        )
        sheet.append(
            (
                "실시간 균열 최소 연결성분 픽셀",
                self.realtime_crack_min_component_pixels,
            )
        )
        sheet.append(
            (
                "레일 구간 정의",
                f"단계별 CAMERA_CAPTURE 수신 순서(1~{RAIL_SECTION_TARGET})",
            )
        )
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 78
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

    def _validate_settings(self) -> None:
        settings = {
            self.settings_sheet.cell(row, 1).value: self.settings_sheet.cell(row, 2).value
            for row in range(2, self.settings_sheet.max_row + 1)
        }
        expected = {
            "파이프라인 버전": PIPELINE_VERSION,
            "Rail Section 목표": RAIL_SECTION_TARGET,
            "캡처 모델 구조": CAPTURE_MODEL_ARCHITECTURE,
            "캡처 모델 파일": self.capture_model_filename,
            "캡처 모델 SHA-256": self.capture_model_sha256,
            "캡처 검출기": self.capture_detector,
            "캡처 입력 전처리": CAPTURE_INPUT_PREPROCESSING,
            "실시간 모델 구조": self.realtime_model_architecture,
            "실시간 모델 파일": self.realtime_model_filename,
            "실시간 모델 SHA-256": self.realtime_model_sha256,
            "실시간 검출기": self.realtime_detector,
            "실시간 입력 전처리": self.realtime_input_preprocessing,
            "캡처 모델 I/O 계약": CAPTURE_MODEL_IO_CONTRACT,
            "실시간 모델 I/O 계약": self.realtime_model_io_contract,
            "결과 기록 정책": REPORT_RESULT_POLICY,
            "클래스 순서": CLASS_ORDER,
            "부식 픽셀 정의": "Fair/Poor/Severe",
            "균열 분석 상태": "enabled" if self.crack_enabled else "disabled",
            "균열 모델 구조": self.capture_crack_model_architecture,
            "균열 모델 파일": self.crack_model_filename,
            "균열 모델 SHA-256": self.crack_model_sha256,
            "균열 검출기": self.crack_detector,
            "균열 입력 전처리": self.capture_crack_input_preprocessing,
            "균열 모델 I/O 계약": self.capture_crack_model_io_contract,
            "균열 확률 임계값": self.crack_probability_threshold,
            "캡처 균열 최소 연결성분 픽셀": (
                self.capture_crack_min_component_pixels
            ),
            "실시간 균열 모델 파일": self.realtime_crack_model_filename,
            "실시간 균열 모델 SHA-256": self.realtime_crack_model_sha256,
            "실시간 균열 검출기": self.realtime_crack_detector,
            "실시간 균열 입력 전처리": self.realtime_crack_input_preprocessing,
            "실시간 균열 모델 I/O 계약": self.realtime_crack_model_io_contract,
            "실시간 균열 확률 임계값": self.realtime_crack_probability_threshold,
            "실시간 균열 최소 연결성분 픽셀": (
                self.realtime_crack_min_component_pixels
            ),
            "레일 구간 정의": (
                f"단계별 CAMERA_CAPTURE 수신 순서(1~{RAIL_SECTION_TARGET})"
            ),
        }
        if settings != expected:
            raise ValueError(
                "Inspection settings differ from the existing report; use a new --report file."
            )

    def _validate_existing_sheet(self) -> None:
        headers = tuple(
            self.sheet.cell(HEADER_ROW, column).value
            for column in range(1, len(HEADERS) + 1)
        )
        if headers != HEADERS:
            raise ValueError("Existing workbook has an incompatible inspection schema.")

        expected_sequence = 1
        phase_counts = {phase: 0 for phase in PHASE_LABELS}
        rescan_started = False
        for row in range(FIRST_DATA_ROW, self.sheet.max_row + 1):
            row_values = [
                self.sheet.cell(row, column).value
                for column in range(1, len(HEADERS) + 1)
            ]
            if all(value is None for value in row_values):
                continue
            sequence = row_values[0]
            if (
                not isinstance(sequence, Integral)
                or isinstance(sequence, bool)
                or int(sequence) != expected_sequence
            ):
                raise ValueError(f"Invalid inspection sequence at row {row}.")
            phase = PHASE_BY_LABEL.get(row_values[1])
            if phase is None:
                raise ValueError(f"Invalid inspection phase at row {row}.")
            if phase == INITIAL_PHASE:
                if rescan_started:
                    raise ValueError(
                        f"Initial rail section follows rescan at row {row}."
                    )
                if phase_counts[INITIAL_PHASE] >= RAIL_SECTION_TARGET:
                    raise ValueError(
                        "Report contains more than "
                        f"{RAIL_SECTION_TARGET} initial rail sections."
                    )
            elif phase == RESCAN_PHASE:
                if phase_counts[INITIAL_PHASE] != RAIL_SECTION_TARGET:
                    raise ValueError(
                        "Rescan starts before "
                        f"{RAIL_SECTION_TARGET} initial rail sections at row {row}."
                    )
                if phase_counts[RESCAN_PHASE] >= RAIL_SECTION_TARGET:
                    raise ValueError(
                        "Report contains more than "
                        f"{RAIL_SECTION_TARGET} rescan rail sections."
                    )
                rescan_started = True
            phase_counts[phase] += 1
            phase_sequence = row_values[2]
            if (
                not isinstance(phase_sequence, Integral)
                or isinstance(phase_sequence, bool)
                or int(phase_sequence) != phase_counts[phase]
            ):
                raise ValueError(f"Invalid phase sequence at row {row}.")
            if str(row_values[5]) != PHASE_TRIGGERS[phase]:
                raise ValueError(f"Invalid phase trigger at row {row}.")
            detector = str(row_values[6])
            if detector != self.capture_detector:
                raise ValueError(
                    f"Unexpected capture detector found at row {row}."
                )
            self._validate_pixels(row_values[7], row_values[8])
            expected_zone = (
                f"{phase_counts[phase]}번 레일 구간"
                if phase in (INITIAL_PHASE, RESCAN_PHASE)
                else "수동"
            )
            if row_values[10] != expected_zone:
                raise ValueError(f"Invalid inspection zone at row {row}.")
            crack_values = self._validate_crack_measurement(
                row_values[11],
                row_values[12],
                row_values[13],
                row_values[14],
                row_values[15],
            )
            expected_crack_formula = (
                f"=IF(P{row}=0,0,O{row}/P{row})"
                if crack_values[0] == "ready"
                else None
            )
            if row_values[16] != expected_crack_formula:
                raise ValueError(f"Invalid crack ratio formula at row {row}.")
            expected_sequence += 1

    def _validate_top_crack_sheet(self) -> None:
        headers = tuple(
            self.top_crack_sheet.cell(TOP_CRACK_HEADER_ROW, column).value
            for column in range(1, len(TOP_CRACK_HEADERS) + 1)
        )
        if headers != TOP_CRACK_HEADERS:
            raise ValueError(
                "Existing workbook has an incompatible top crack schema."
            )

        expected_sequence = 1
        phase_counts = {phase: 0 for phase in PHASE_LABELS}
        rescan_started = False
        for row in range(
            TOP_CRACK_FIRST_DATA_ROW,
            self.top_crack_sheet.max_row + 1,
        ):
            row_values = [
                self.top_crack_sheet.cell(row, column).value
                for column in range(1, len(TOP_CRACK_HEADERS) + 1)
            ]
            if all(value is None for value in row_values):
                continue
            sequence = row_values[0]
            if (
                not isinstance(sequence, Integral)
                or isinstance(sequence, bool)
                or int(sequence) != expected_sequence
            ):
                raise ValueError(f"Invalid top crack sequence at row {row}.")
            phase = PHASE_BY_LABEL.get(row_values[1])
            if phase is None:
                raise ValueError(f"Invalid top crack phase at row {row}.")
            phase_counts[phase] += 1
            phase_sequence = row_values[2]
            if phase in (INITIAL_PHASE, RESCAN_PHASE):
                if phase == INITIAL_PHASE:
                    if rescan_started:
                        raise ValueError(
                            f"Top initial rail section follows rescan at row {row}."
                        )
                else:
                    if phase_counts[INITIAL_PHASE] != RAIL_SECTION_TARGET:
                        raise ValueError(
                            "Top rescan starts before "
                            f"{RAIL_SECTION_TARGET} initial rail sections "
                            f"at row {row}."
                        )
                    rescan_started = True
                if phase_counts[phase] > RAIL_SECTION_TARGET:
                    raise ValueError(
                        "Top report contains more than "
                        f"{RAIL_SECTION_TARGET} {phase} rail sections."
                    )
                if (
                    not isinstance(phase_sequence, Integral)
                    or isinstance(phase_sequence, bool)
                    or int(phase_sequence) != phase_counts[phase]
                ):
                    raise ValueError(
                        f"Invalid top phase sequence at row {row}."
                    )
            elif phase_sequence is not None:
                raise ValueError(f"Manual top sequence must be empty at row {row}.")
            if str(row_values[6]) != PHASE_TRIGGERS[phase]:
                raise ValueError(f"Invalid top capture trigger at row {row}.")
            if row_values[7] != "top":
                raise ValueError(f"Invalid top camera role at row {row}.")
            crack_values = self._validate_crack_measurement(
                row_values[8],
                row_values[9],
                row_values[10],
                row_values[11],
                row_values[12],
            )
            if crack_values[0] != "ready":
                raise ValueError(f"Top crack analysis is not ready at row {row}.")
            expected_formula = f"=IF(M{row}=0,0,L{row}/M{row})"
            if row_values[13] != expected_formula:
                raise ValueError(f"Invalid top crack ratio formula at row {row}.")
            expected_zone = (
                f"{phase_sequence}번 레일 구간"
                if phase in (INITIAL_PHASE, RESCAN_PHASE)
                else "수동"
            )
            if row_values[14] != expected_zone:
                raise ValueError(f"Invalid top inspection zone at row {row}.")
            captured_at = self._normalize_captured_at(row_values[3])
            if not self._matching_side_capture_exists(
                phase,
                int(phase_sequence) if phase_sequence is not None else None,
                captured_at,
            ):
                raise ValueError(
                    f"Top crack row {row} has no matching SIDE capture."
                )
            expected_sequence += 1

    def _initialize_sheet(self, sheet: Worksheet) -> None:
        sheet.merge_cells("A1:Q1")
        title = sheet["A1"]
        title.value = "표면 부식·균열 분할 점검 보고서"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor="1F4E78")
        title.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

        for column, label in enumerate(
            (
                "검사 단계",
                "Rail Section 수",
                "부식 픽셀 합",
                "검사 픽셀 합",
                "전체 부식률(픽셀 가중)",
            ),
            start=1,
        ):
            cell = sheet.cell(row=2, column=column, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center")
            cell.border = CELL_BORDER

        sheet["A3"] = PHASE_LABELS[INITIAL_PHASE]
        sheet["A4"] = PHASE_LABELS[RESCAN_PHASE]
        self._update_summary_formulas(FIRST_DATA_ROW)
        for row in (3, 4):
            for column in range(1, 6):
                cell = sheet.cell(row=row, column=column)
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
                cell.alignment = Alignment(horizontal="right" if column > 1 else "center")
                cell.border = CELL_BORDER
            for column in (2, 3, 4):
                sheet.cell(row=row, column=column).number_format = "#,##0"
            sheet.cell(row=row, column=5).number_format = "0.00%"

        sheet["G2"] = "부식률 감소량"
        sheet["G3"] = "초기 대비 개선율"
        sheet["G4"] = "개선율 0분모"
        sheet.merge_cells("H4:J4")
        sheet["H4"] = "초기·재검사 모두 0%면 0%, 초기만 0%면 빈칸"
        for coordinate in ("G2", "G3", "G4"):
            cell = sheet[coordinate]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center")
            cell.border = CELL_BORDER
        for coordinate in ("H2", "H3", "H4"):
            cell = sheet[coordinate]
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
            cell.alignment = Alignment(horizontal="right")
            cell.border = CELL_BORDER
        for coordinate in ("H2", "H3"):
            sheet[coordinate].number_format = "0.00%"
        sheet["H4"].alignment = Alignment(horizontal="left", wrap_text=True)
        sheet.row_dimensions[4].height = 30

        for column, heading in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=HEADER_ROW, column=column, value=heading)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER

        widths = (
            8,
            14,
            11,
            21,
            52,
            18,
            26,
            16,
            16,
            16,
            14,
            16,
            34,
            16,
            16,
            18,
            16,
        )
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + column)].width = width
        sheet.freeze_panes = f"A{FIRST_DATA_ROW}"
        sheet.auto_filter.ref = f"A{HEADER_ROW}:Q{HEADER_ROW}"
        sheet.sheet_view.showGridLines = False

    def _initialize_top_crack_sheet(self, sheet: Worksheet) -> None:
        for column, heading in enumerate(TOP_CRACK_HEADERS, start=1):
            cell = sheet.cell(
                row=TOP_CRACK_HEADER_ROW,
                column=column,
                value=heading,
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7030A0")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER
        widths = (8, 14, 11, 21, 52, 52, 18, 14, 16, 34, 16, 16, 18, 16, 14)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + column)].width = width
        sheet.freeze_panes = f"A{TOP_CRACK_FIRST_DATA_ROW}"
        sheet.auto_filter.ref = (
            f"A{TOP_CRACK_HEADER_ROW}:O{TOP_CRACK_HEADER_ROW}"
        )
        sheet.sheet_view.showGridLines = False

    def _next_data_row(self) -> int:
        for row in range(self.sheet.max_row, FIRST_DATA_ROW - 1, -1):
            if self.sheet.cell(row=row, column=1).value is not None:
                return row + 1
        return FIRST_DATA_ROW

    def _next_top_crack_row(self) -> int:
        for row in range(
            self.top_crack_sheet.max_row,
            TOP_CRACK_FIRST_DATA_ROW - 1,
            -1,
        ):
            if self.top_crack_sheet.cell(row=row, column=1).value is not None:
                return row + 1
        return TOP_CRACK_FIRST_DATA_ROW

    def _format_data_row(self, row: int) -> None:
        for column in range(1, len(HEADERS) + 1):
            cell = self.sheet.cell(row=row, column=column)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="left" if column in (5, 6, 7, 11, 12, 13) else "center",
                vertical="center",
            )
        self.sheet.cell(row, 4).number_format = "yyyy-mm-dd hh:mm:ss"
        self.sheet.cell(row, 8).number_format = "#,##0"
        self.sheet.cell(row, 9).number_format = "#,##0"
        self.sheet.cell(row, 10).number_format = "0.00%"
        self.sheet.cell(row, 15).number_format = "#,##0"
        self.sheet.cell(row, 16).number_format = "#,##0"
        self.sheet.cell(row, 17).number_format = "0.000%"

    def _format_top_crack_row(self, row: int) -> None:
        for column in range(1, len(TOP_CRACK_HEADERS) + 1):
            cell = self.top_crack_sheet.cell(row=row, column=column)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column in (5, 6, 7, 8, 9, 10, 15)
                    else "center"
                ),
                vertical="center",
            )
        self.top_crack_sheet.cell(row, 4).number_format = "yyyy-mm-dd hh:mm:ss"
        self.top_crack_sheet.cell(row, 12).number_format = "#,##0"
        self.top_crack_sheet.cell(row, 13).number_format = "#,##0"
        self.top_crack_sheet.cell(row, 14).number_format = "0.000%"

    def _update_summary_formulas(self, last_row: int) -> None:
        for row in (3, 4):
            self.sheet[f"B{row}"] = (
                f'=COUNTIF(B{FIRST_DATA_ROW}:B{last_row},A{row})'
            )
            self.sheet[f"C{row}"] = (
                f'=SUMIF(B{FIRST_DATA_ROW}:B{last_row},A{row},'
                f'H{FIRST_DATA_ROW}:H{last_row})'
            )
            self.sheet[f"D{row}"] = (
                f'=SUMIF(B{FIRST_DATA_ROW}:B{last_row},A{row},'
                f'I{FIRST_DATA_ROW}:I{last_row})'
            )
            self.sheet[f"E{row}"] = f"=IF(D{row}=0,0,C{row}/D{row})"
        self.sheet["H2"] = (
            f'=IF(OR(B3<>{RAIL_SECTION_TARGET},B4<>{RAIL_SECTION_TARGET}),"",E3-E4)'
        )
        self.sheet["H3"] = (
            f'=IF(OR(B3<>{RAIL_SECTION_TARGET},B4<>{RAIL_SECTION_TARGET}),"",'
            'IF(E3=0,IF(E4=0,0,""),(E3-E4)/E3))'
        )

    def _phase_counts(self) -> dict[str, int]:
        counts = {phase: 0 for phase in PHASE_LABELS}
        for row in range(FIRST_DATA_ROW, self._next_data_row()):
            phase = PHASE_BY_LABEL.get(self.sheet.cell(row, 2).value)
            if phase is not None:
                counts[phase] += 1
        return counts

    def _top_crack_phase_counts(self) -> dict[str, int]:
        counts = {phase: 0 for phase in PHASE_LABELS}
        for row in range(
            TOP_CRACK_FIRST_DATA_ROW,
            self._next_top_crack_row(),
        ):
            phase = PHASE_BY_LABEL.get(self.top_crack_sheet.cell(row, 2).value)
            if phase is not None:
                counts[phase] += 1
        return counts

    def _matching_side_capture_exists(
        self,
        phase: str,
        phase_sequence: int | None,
        captured_at: datetime,
    ) -> bool:
        phase_label = PHASE_LABELS[phase]
        for row in range(FIRST_DATA_ROW, self._next_data_row()):
            if self.sheet.cell(row, 2).value != phase_label:
                continue
            if phase in (INITIAL_PHASE, RESCAN_PHASE):
                if self.sheet.cell(row, 3).value == phase_sequence:
                    return True
                continue
            side_captured_at = self.sheet.cell(row, 4).value
            if isinstance(side_captured_at, datetime) and abs(
                (side_captured_at - captured_at).total_seconds()
            ) <= 0.001:
                return True
        return False

    def _validate_crack_measurement(
        self,
        status: str,
        detector: str | None,
        detected: bool | None,
        crack_pixels: int | None,
        inspected_pixels: int | None,
    ) -> tuple[str, str | None, bool | None, int | None, int | None]:
        status = str(status)
        expected_status = "ready" if self.crack_enabled else "disabled"
        if status != expected_status:
            raise ValueError(
                f"crack_status must be {expected_status!r} for this report."
            )
        if status == "disabled":
            if any(
                value is not None
                for value in (
                    detector,
                    detected,
                    crack_pixels,
                    inspected_pixels,
                )
            ):
                raise ValueError(
                    "Disabled crack analysis must not contain detector or pixel data."
                )
            return status, None, None, None, None

        detector = str(detector)
        if detector != self.crack_detector:
            raise ValueError(
                "Only results from this report's exact crack detector can be added."
            )
        if not isinstance(detected, bool):
            raise ValueError("crack_detected must be a boolean for ready results.")
        crack_pixels, inspected_pixels = self._validate_measurement_pixels(
            crack_pixels,
            inspected_pixels,
            "crack",
        )
        if detected != (crack_pixels > 0):
            raise ValueError(
                "crack_detected must agree with whether crack_pixels is positive."
            )
        return status, detector, detected, crack_pixels, inspected_pixels

    @staticmethod
    def _validate_pixels(rust_pixels: int, inspected_pixels: int) -> tuple[int, int]:
        return InspectionWorkbook._validate_measurement_pixels(
            rust_pixels,
            inspected_pixels,
            "rust",
        )

    @staticmethod
    def _validate_measurement_pixels(
        positive_pixels: int,
        inspected_pixels: int,
        label: str,
    ) -> tuple[int, int]:
        if (
            not isinstance(positive_pixels, Integral)
            or isinstance(positive_pixels, bool)
            or not isinstance(inspected_pixels, Integral)
            or isinstance(inspected_pixels, bool)
        ):
            raise ValueError("Pixel counts must be integers.")
        positive_pixels = int(positive_pixels)
        inspected_pixels = int(inspected_pixels)
        if positive_pixels < 0:
            raise ValueError(f"{label}_pixels must not be negative.")
        if inspected_pixels <= 0:
            raise ValueError("inspected_pixels must be greater than zero.")
        if positive_pixels > inspected_pixels:
            raise ValueError(f"{label}_pixels must not exceed inspected_pixels.")
        return positive_pixels, inspected_pixels

    @staticmethod
    def _validate_model_sha256(model_sha256: str) -> str:
        model_sha256 = str(model_sha256).lower()
        if len(model_sha256) != 64 or any(
            character not in "0123456789abcdef" for character in model_sha256
        ):
            raise ValueError("model_sha256 must be a 64-character hexadecimal digest.")
        return model_sha256

    @staticmethod
    def _normalize_captured_at(captured_at: datetime | None) -> datetime:
        if captured_at is None:
            return datetime.now()
        if not isinstance(captured_at, datetime):
            raise TypeError("captured_at must be a datetime or None.")
        if captured_at.tzinfo is not None:
            return captured_at.astimezone().replace(tzinfo=None)
        return captured_at
